
import re
import requests
import json
import openpyxl
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


TOKEN = "pat"
API_URL = "https://stib.net"

names = [
    ["Last update","lastFetch"],
    ["Id","id"],
    ["State","state"],
    ["Start","startDate"],
    ["End","endDate"],
    ["Name","name"],
    ["Issues At Start","issuesAtStart"],
    ["Issues Done","issuesDone"],
    ["Stories At Start","storiesAtStart"],
    ["Stories Done","storiesDone"],
    ["Bugs total","bugs"],
    ["bugs Added","bugsAdded"],
    ["SP at Start","storyPointsAtStart"],
    ["SP Done","storyPointsDone"],
    ["Issues Not Done","issuesNotDone"],
    ["issues Added","issuesAdded"],
    ["Issues Removed","issuesRemoved"],
     ["Stories Not Done","storiesNotDone"],
    ["Stories Added",   "storiesAdded"],
    ["Stories Removed", "storiesRemoved"],
    ["SP Not Done","storyPointsNotDone"],
    ["SP Added","storyPointsAdded"],
    ["Points at Start","pointsAtStart"],
    ["Points Done","pointsDone"],
    ["Points Not Done","pointsNotDone"],
    ["Points Added","pointsAdded"],
    ["Points Removed","pointsRemoved"],
    ["Active Epics","activeEpics"]
   
 
]      

def getJQL(jql,fields):
    query_params = {
        "jql": jql,
        "fields": fields  
        }
    return getFromAPI("/rest/api/2/search",query_params)

def getFromAPI(path,query_params = None):
        url = API_URL + path
        headers = {
            "Authorization": "Basic " + TOKEN,
            "Content-Type": "application/json"
        }
        if query_params is None:
            resp = requests.get(url, headers=headers)
        else:
            resp = requests.get(url, headers=headers,params=query_params)
    
        if resp.status_code != 200:
            print("Error retrieving data for url {}: {}".format(url, resp.text))
            return ""
        else:
            return resp.text
    

logs =[]
def addLog(msg):
    logs.append(msg)
    print(msg)


def match_sprint(  text):
    # Extracting year, sequence, and name separately
    year_pattern = r"(\d{4})"
    sequence_pattern = r"S(\d+)"
    name_pattern = r"-(\w+)$"

    matched_year = re.search(year_pattern, text).group() if re.search(year_pattern, text) else None
    matched_seq = re.search(sequence_pattern, text).group(1).lstrip('0') if re.search(sequence_pattern, text) else None
    matched_name = re.search(name_pattern, text).group(1) if re.search(name_pattern, text) else None
    
    # Validating the extracted values against the provided inputs
    return matched_year, matched_seq, matched_name
   

class DictToClass:
    def __init__(self, ji):
        self.__dict__ = ji
        
class SprintReport(dict):
    def __init__(self, sprintr):
        #ref of fields
        dc = DictToClass(sprintr)
        self.id=dc.id
        self.name = dc.name
        self.goal = sprintr.get("goal", "")
        self.state= dc.state
        self.startDate =datetime.strptime(dc.startDate , "%Y-%m-%dT%H:%M:%S.%f%z").strftime("%d/%m/%Y %H:%M:%S")
        self.endDate = datetime.strptime(dc.endDate , "%Y-%m-%dT%H:%M:%S.%f%z").strftime("%d/%m/%Y %H:%M:%S")
        #from the sprint report 
        self.hasReportData = False
        self.issuesAdded = 0
        self.storiesAdded = 0
        self.bugsAdded = 0
        self.pointsAdded = 0
        
        self.storiesDone = 0
        self.issuesDone = 0
        self.pointsDone = 0
        self.storiesNotDone = 0
        self.storyPointsNotDone = 0
        self.storyPointsDone = 0
        self.storyPointsAdded =0
        self.issuesNotDone = 0
        self.pointsNotDone = 0

        self.issuesRemoved = 0
        self.storiesRemoved = 0
        self.pointsRemoved = 0
        self.storiesAtStart = 0
        self.issuesAtStart = 0
        self.pointsAtStart = 0
        self.activeEpics = 0
        self.bugs = 0

        

        #copy sprint info from sprint dict field
        #self.__dict__ = sprintr.sprint
        #contents = DictToClass(sprintr.contents)
    def getPoints(sef, issue):
        if 'estimateStatistic' in issue:
            if 'statFieldValue' in issue['estimateStatistic']:
                if 'value' in issue['estimateStatistic']['statFieldValue']:
                    return int(issue['estimateStatistic']['statFieldValue']['value'])
        return 0

    def updateFromJiraReport(self,report):
        self.lastFetch =  datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        c = report['contents']
        
        addedKeys=[]
        if 'issueKeysAddedDuringSprint' in c:
            addedKeys = c['issueKeysAddedDuringSprint']

        for issue in c['completedIssues']:
            points = self.getPoints(issue)
            self.issuesDone += 1
            self.pointsDone += points
            if issue['key'] in addedKeys:
                self.issuesAdded += 1
                self.pointsAdded += points

            if issue["typeId"] == "7": # is it a story ?
                self.storiesDone += 1
                self.storyPointsDone += points
                if issue['key'] in addedKeys:
                    self.storiesAdded += 1
            elif issue["typeId"] == "1": # is it a bug ?
                self.bugs +=1
                if issue['key'] in addedKeys:
                    self.bugsAdded += 1

                
                      
        for issue in c['issuesNotCompletedInCurrentSprint']:
            self.issuesNotDone += 1
            points = self.getPoints(issue)
            self.pointsNotDone += points
            if issue['key'] in addedKeys:
                self.issuesAdded +=1
                self.pointsAdded += points

            if issue["typeId"] == "7": # is it a story ?
                self.storiesNotDone += 1
                self.storyPointsNotDone += points
                if issue['key'] in addedKeys:
                    self.storiesAdded += 1
                    self.storyPointsAdded += points
            elif issue["typeId"] == "1": # is it a bug ?
                self.bugs +=1
                if issue['key'] in addedKeys:
                    self.bugsAdded += 1
          
                
                
      
        if 'puntedIssues' in c:
            self.issuesRemoved = len(c['puntedIssues'])
            for issue in c['puntedIssues']:
                self.pointsRemoved += self.getPoints(issue)
                if issue["typeId"] == "7": # is it a story ?
                    self.storiesRemoved += 1
        
        self.issuesAtStart = self.issuesDone + self.issuesNotDone - self.issuesAdded + self.issuesRemoved
        self.storiesAtStart = self.storiesDone + self.storiesNotDone - self.storiesAdded + self.storiesRemoved
        self.pointsAtStart = self.pointsDone + self.pointsNotDone - self.pointsAdded + self.pointsRemoved
        self.storyPointsAtStart = self.storyPointsDone + self.storyPointsNotDone 

       
  

def getBoardsForTeams(wb):
    try:
       
        sheet = wb['Setup']
        startRow = int(sheet.cell(row=3, column=2).value)
        endRow = int(sheet.cell(row=3, column=3).value)
       
        boards = []
        for row in range(startRow, endRow + 1):
            team = sheet.cell(row=row, column=2).value
            board = int(sheet.cell(row=row, column=3).value)
            skip = int(sheet.cell(row=row, column=4).value)
            boards.append(DictToClass({"team": team, "id": board, "skip": skip}))
        return boards
        wb()
    except Exception as e:
        print(e) 

def getEpics(sprintReport):
    #print(f"  ..Getting Epics for sprint {sprintReport.id} : {sprintReport.name}")
    sprintItems = json.loads(getJQL(f"sprint = {sprintReport.id} and issuetype not in ( Sub-task, Bug)",fields = "key,summary,customfield_10006")
                             )
    epicSet = set()
    for item in sprintItems["issues"]:
        epicSet.add(item["fields"]["customfield_10006"])

    return len(epicSet)



#get the sprints
print("______________________________________________________________________________")
print("Getting Sprints for boards")
print("______________________________________________________________________________")

wb = load_workbook('RVSprints.xlsx')
boards = getBoardsForTeams(wb) 

resultcount = 0
startAt = 0
startRow = 6
numRows = len(names);
allTeams = ", ".join([obj.team for obj in boards])

sprintboards=[]

# for each team (or board)
for nboard in range(len(boards)):
    
    skip = boards[nboard].skip  # if we know that we can skip the first - old sprints, skip will do that
    col = 2
    row = startRow + nboard * numRows
    moreResults = True

    addLog(f"Getting sprints for {boards[nboard].team} , skipping first {skip} sprints")
 
    currentTime = datetime.now()
    url = "/rest/agile/1.0/board/" + str(boards[nboard].id) + "/sprint?startAt=" + str(skip) + "&state=active,closed"
    rtext = getFromAPI(url)
    result = json.loads(rtext)
    sprintboard = DictToClass({ "board": boards[nboard],  "sprints":[]})
    
    while moreResults:
        for i in range(len(result["values"])):
            sprint = result["values"][i]
            sr = SprintReport(sprint)
            
            if boards[nboard].team in sprint["name"]:
                sprintboard.sprints.append(sr)
            skip += 1

        moreResults = False
        if len(result["values"]) > 49:
            addLog(f"Getting sprints for {boards[nboard].team} , skipping first {skip} sprints")
            addLog(f"  skip can be increased for {boards[nboard].team}")
 
            moreResults = True
            result = json.loads(getFromAPI("/rest/agile/1.0/board/" + str(boards[nboard].id) + "/sprint?startAt=" + str(skip) + "&state=active,closed"))

    sprintboards.append(sprintboard)


#only keep the sprints we want 
#load the sprints we want to have from the first row int the teams sheet
neededSprints = []
sheet = wb['Teams']
for col in range(2,20):
    sprint = sheet.cell(row=1, column=col).value
    if sprint : 
        neededSprints.append(DictToClass({"sprint": sprint, "col":col}))
        


sb2 = []

#filter the boards to keep only sprints from that team and the sprint number from excel 
for sb in sprintboards:
    board=DictToClass({"board":sb.board,"sprints":[]})
    print(f"Filtering sprints from board {sb.board.team}")
    keepSprint = False
    for sprint in sb.sprints:
        sprintname = sprint.name
        mYear,mSprint,mTeam = match_sprint(sprint.name)
        for t in neededSprints:
            tYear,tSprint,tTeam = match_sprint(t.sprint)
            if mYear==tYear and mSprint==tSprint:
                keepSprint = True
        if keepSprint:
            board.sprints.append(sprint)
            print(f"  ..Keeping Sprint {sprint.id} - {sprint.name}, found year : {mYear}, Sprint sequence : {mSprint}, Team : {mTeam}")
            keepSprint= False
    sb2.append(board)
        

sprintboards = sb2
#now get the sprint reports for the sprints we need 
for sb in sprintboards:
    sprints = sb.sprints
    for sprint in sprints:
        url = f"/rest/greenhopper/latest/rapid/charts/sprintreport?rapidViewId={ sb.board.id }&sprintId={sprint.id}"
        report = json.loads( getFromAPI(url ) )
        print(f"Updating sprint {report['sprint']['name']}")
        sprint.updateFromJiraReport(report)
        numEpics = getEpics(sprint)
        sprint.activeEpics = numEpics

#saving to excel
print('Writing data to Teams sheet')
#wb = load_workbook(filename="wlsprints.xlsm")
sheet = wb['Teams']
sheet.delete_rows(3, 1000)
if 'Data' not in wb.sheetnames:
    wb.create_sheet('Data')
sheet2 = wb['Data']
sheet2.delete_rows(1, 1000)



col = 1
r = 3
for sb in sprintboards: #enumerating teams , eaxh team in a separate row
    col= 1
    print(f"Writing Excel sprints for team {sb.board.team}")
    sheet.cell(column=col,row=r, value=sb.board.team)
    r += 1
    startRow = r
    maxCol = 2
    namesrow = r
    for name in names: #fill first column wirh names
        sheet.cell(column=col,row=namesrow, value=name[0])
        namesrow+=1
    
    col=2
    for sprint in sb.sprints:
        #finding the right column
        for ns in neededSprints:
            if ns.sprint in sprint.name:
                col = ns.col
                maxCol = col if col > maxCol else col
                

        s=vars(sprint)
        namesrow=r
        for name in names:
            n=name[1]
            try:
                if n :
                    sheet.cell(column=col,row=namesrow, value=s[n])
            except Exception as e:
                print(e)
            namesrow+=1
        col +=1
    
    #ready to go to next team
    r+=len(names)+1

print("saving")
try:
    wb.save("wlsprints.xlsx")
except  Exception as e :
    print(f"I was unable to save the excel wlsprints.xlsx .. is it open? {e} ") 

print("done")
# input("Type a key to finish and close the graphs ")

#writing to sxl:
            # specific for worldline
            
             
            #col += 1
            #sheet.cell(row=row, column=2, values=names)
            #sheet.cell(row=row, column=1).value = boards[nboard]["team"]
            #sheet.cell(row=row+1, column=1).value = boards[nboard]["board"]

            # values = [
            #     [currentTime.strftime("%Y-%m-%d %H:%M:%S")],
            #     [sr.id],
            #     [sr.state],
            #     [sr.startDate],
            #     [sr.endDate],
            #     [sr.name]
            # ]
            # for r, row_values in enumerate(values, start=row):
            #     for c, value in enumerate(row_values, start=col):
            #         sheet.cell(row=r, column=c).value = value
    


